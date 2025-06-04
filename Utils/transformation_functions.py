
import tempfile
import shutil
import pandas as pd
from typing import List
from loguru import logger
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, column_index_from_string


def seleccionar_columnas_pd(
    df: pd.DataFrame, cols_elegidas: List[str]
) -> pd.DataFrame | None:
    """
    Filtra y retorna las columnas especificadas del DataFrame.

    Parámetros:
    dataframe (pd.DataFrame): DataFrame del cual se filtrarán las columnas.
    cols_elegidas (list): Lista de nombres de las columnas a incluir en el DataFrame filtrado.

    Retorna:
    pd.DataFrame: DataFrame con las columnas filtradas.
    """
    try:
        # Verificar si dataframe es un DataFrame de pandas
        if not isinstance(df, pd.DataFrame):
            raise TypeError("El argumento 'dataframe' debe ser un DataFrame de pandas.")

        # Filtrar las columnas especificadas
        df_filtrado = df[cols_elegidas]

        # Registrar el proceso
        #logger.info(f"Columnas filtradas: {', '.join(cols_elegidas)}")

        return df_filtrado

    except KeyError as ke:
        error_message = (
            f"Error: Columnas especificadas no encontradas en el DataFrame: {str(ke)}"
        )
        logger.error(error_message)
        return df
    except Exception as e:
        logger.critical(f"Error inesperado al filtrar columnas: {str(e)}")


def pd_left_merge(
    base_left: pd.DataFrame, base_right: pd.DataFrame, key: str
) -> pd.DataFrame:
    """Función que retorna el left join de dos dataframe de pandas.

    Args:
        base_left (pd.DataFrame): Dataframe que será la base del join.
        base_right (pd.DataFrame): Dataframe del cuál se extraerá la información    complementaria.
        key (str): Llave mediante la cual se va a realizar el merge o join.

    Returns:
        pd.DataFrame: Dataframe con el merge de las dos fuentes de datos.
    """

    # Validar que base_left y base_right sean DataFrames de pandas
    if not isinstance(base_left, (pd.DataFrame, pd.Series)):
        raise ValueError("El argumento base_left no es un DataFrame de pandas")
    if not isinstance(base_right, (pd.DataFrame, pd.Series)):
        raise ValueError("El argumento base_right no es un DataFrame de pandas")

    base = None

    try:
        base = pd.merge(left=base_left, right=base_right, how="left", on=key)
        logger.success("Proceso de merge satisfactorio")
    except pd.errors.MergeError as e:
        logger.critical(f"Proceso de merge fallido: {e}")
        raise e

    return base


def pd_left_merge_two_keys(
    base_left: pd.DataFrame,
    base_right: pd.DataFrame,
    left_key: str,
    right_key: str,
) -> pd.DataFrame:
    """Función que retorna el left join de dos dataframe de pandas.

    Args:
        base_left (pd.DataFrame): Dataframe que será la base del join.
        base_right (pd.DataFrame): Dataframe del cuál se extraerá la información complementaria.
        key (str): Llave mediante la cual se va a realizar el merge o join.

    Returns:
        pd.DataFrame: Dataframe con el merge de las dos fuentes de datos.
    """

    # Validar que base_left y base_right sean DataFrames de pandas
    if not isinstance(base_left, (pd.DataFrame, pd.Series)):
        raise ValueError("El argumento base_left no es un DataFrame de pandas")
    if not isinstance(base_right, (pd.DataFrame, pd.Series)):
        raise ValueError("El argumento base_right no es un DataFrame de pandas")

    base = None

    try:
        base = pd.merge(
            left=base_left,
            right=base_right,
            how="left",
            left_on=left_key,
            right_on=right_key,
        )
        logger.success("Proceso de merge satisfactorio")
    except pd.errors.MergeError as e:
        logger.critical(f"Proceso de merge fallido: {e}")
        raise e

    return base



class ExcelPlantilla:
    """
    Clase que permite trabajar con una plantilla de Excel, insertando datos en ella y
    generando archivos de salida con funcionalidades como listas desplegables.
    """

    COL_MOTIVOS = "E"  # Columna predeterminada para aplicar motivos de devolución

    def __init__(self, wb=None, hoja="Hoja 1", ruta_plantilla=None):
        """
        Inicializa una instancia de ExcelPlantilla.

        Args:
            wb (Workbook, opcional): Libro de trabajo de openpyxl ya cargado.
            hoja (str): Nombre de la hoja sobre la cual trabajar.
            ruta_plantilla (str, opcional): Ruta al archivo base .xlsx, útil para clonación.
        """
        self.hoja = hoja
        self.ruta_plantilla = ruta_plantilla
        self.ruta_salida = None

        if wb:
            self.wb = wb
            self.ws = self.wb[self.hoja]

    @classmethod
    def cargar_desde_archivo(cls, ruta_plantilla: str, hoja: str = "Hoja 1"):
        """
        Crea una instancia base con la ruta de la plantilla sin cargar el archivo aún.

        Args:
            ruta_plantilla (str): Ruta al archivo .xlsx que actúa como plantilla base.
            hoja (str): Hoja de Excel donde se escribirán los datos.

        Returns:
            ExcelPlantilla: Instancia preparada para clonar.
        """
        return cls(ruta_plantilla=ruta_plantilla, hoja=hoja)

    def clonar_con_salida(self, ruta_salida: str):
        """
        Crea una copia física del archivo de plantilla en disco, la abre y actualiza
        la instancia actual para trabajar sobre la copia (no una nueva instancia).

        Args:
            ruta_salida (str): Ruta donde se guardará el nuevo archivo generado.

        Returns:
            self: La misma instancia actual con ruta_salida y libro/hoja actualizados.
        """
        if not self.ruta_plantilla:
            raise ValueError("No se ha definido ruta_plantilla en la instancia base.")

        # Copiar físicamente el archivo
        shutil.copyfile(self.ruta_plantilla, ruta_salida)

        # Cargar el nuevo archivo y actualizar atributos
        self.ruta_salida = ruta_salida
        self.wb = load_workbook(ruta_salida)
        self.ws = self.wb[self.hoja]

        return self


    def insertar_dataframe(self, df: pd.DataFrame):
        """
        Inserta un DataFrame en la hoja activa, comenzando desde la siguiente fila disponible.

        Args:
            df (pd.DataFrame): DataFrame con los datos a insertar. No se incluye encabezado.
        """
        for fila in dataframe_to_rows(df, index=False, header=False):
            self.ws.append(fila)

    def aplicar_lista_desplegable(self, columna: str, opciones: list):
        """
        Aplica una validación tipo lista desplegable en toda la columna especificada,
        comenzando desde la fila 5 hasta la última fila con datos.

        Args:
            columna (str): Letra de la columna (ej. "E") donde aplicar la lista desplegable.
            opciones (list): Lista de opciones visibles en el desplegable.

        Raises:
            Exception: Si ocurre un error durante la aplicación de la validación.
        """
        try:
            col_num = column_index_from_string(columna.upper())
            dv = DataValidation(
                type="list",
                formula1=f'"{",".join(opciones)}"',
                allow_blank=True
            )
            self.ws.add_data_validation(dv)

            col_letra = get_column_letter(col_num)
            rango = f"{col_letra}5:{col_letra}{self.ws.max_row}"
            dv.add(rango)
        except Exception as e:
            logger.error(f"Error aplicando validación: {e}")
            raise

    def guardar(self):
        """
        Guarda el archivo en la ruta de salida definida.

        Raises:
            ValueError: Si no se ha definido ruta_salida antes de guardar.
        """
        if not self.ruta_salida:
            raise ValueError("Debe establecerse una ruta_salida antes de guardar.")
        self.wb.save(self.ruta_salida)

def Crear_diccionario_desde_dataframe(
        df: pd.DataFrame, col_clave: str, col_valor: str
    ) -> dict:
        """
        Crea un diccionario a partir de un DataFrame utilizando dos columnas especificadas.

        Args:
            df (pd.DataFrame): El DataFrame de entrada.
            col_clave (str): El nombre de la columna que se utilizará como clave en el diccionario.
            col_valor (str): El nombre de la columna que se utilizará como valor en el diccionario.

        Returns:
            dict: Un diccionario creado a partir de las columnas especificadas.
        """
        try:
            # Verificar si las columnas existen en el DataFrame
            if col_clave not in df.columns or col_valor not in df.columns:
                raise ValueError(
                    "Las columnas especificadas no existen en el DataFrame."
                )

            # Crear el diccionario a partir de las columnas especificadas
            resultado_dict = df.set_index(col_clave)[col_valor].to_dict()

            return resultado_dict

        except ValueError as ve:
            # Registrar un mensaje crítico si hay un error
            logger.critical(f"Error: {ve}")
            raise ve